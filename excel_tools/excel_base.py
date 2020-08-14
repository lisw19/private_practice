from openpyxl import load_workbook


class ExcelUser(object):
    """
    excel 增、删、改、查基础操作
    """

    def __init__(self, excel_name):
        """
        请输入需要操作的Excel路径（默认当前目录下）
        """
        self._excel_name = excel_name
        self._wb = load_workbook(self._excel_name)

    def clean(self):
        """
        清除除表头外所有数据
        """
        sheet_list = self._wb.get_sheet_names()
        for each in sheet_list:
            ws = self._wb.get_sheet_by_name(each)
            for i in range(2, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    ws.cell(row=i, column=j).value = ""
        return None

    def clean_but(self, unclean_list):
        """
        清除除指定sheet外所有数据
        :param unclean_list:
        :return
        """
        sheet_list = self._wb.get_sheet_names()
        for each in sheet_list:
            if each in unclean_list:
                continue
            ws = self._wb.get_sheet_by_name(each)
            for i in range(2, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    ws.cell(row=i, column=j).value = ""
        return None

    def read(self, sheet_name):
        """
        读取数据
        :param sheet_name:
        :return:
        """
        data = []
        ws = self._wb.get_sheet_by_name(sheet_name)
        for i in range(1, ws.max_row + 1):
            temp = []
            for j in range(1, ws.max_column + 1):
                temp.append(ws.cell(row=i, column=j).value)
            data.append(temp)
        return data

    def insert(self, sheet_name, data, begin_row=3):
        """
        插入表格数据
        :param sheet_name:
        :param data:想要插入的数据（二维数组 ）
        :param begin_row: 从第几行开始插入
        :return:
        """
        if data == () or len(data[0]) is None:
            return
        ws = self._wb.get_sheet_by_name(sheet_name)
        for i in range(begin_row, len(data) + begin_row):
            for j in range(1, len(data[0]) + 1):
                ws.cell(row=i, column=j).value = data[i - begin_row][j - 1]
        return

    def append(self, sheet_name, data_list):
        """
        在表的底部追加
        :param sheet_name: 表格的sheet
        :param data_list: 追加的数据 type：list
        :return:
        """
        ws = self._wb.get_sheet_by_name(sheet_name)
        ROW = ws.max_row + 1
        for i in range(1, len(data_list) + 1):
            ws.cell(row=ROW, column=i).value = data_list[i - 1]
        return None

    def save(self, name=None):
        """
        保存
        :param name:
        :return:
        """
        if name is None:
            self._wb.save(self._excel_name)
        else:
            self._wb.save(name)
        print("保存成功")

    def save_as(self, name):
        self.save(name)
        return
