# -*- coding: utf-8 -*-
import time
import sys
import glob

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from excel_tools.excel_base import ExcelUser
from tools.time_package import get_first_day_of_month, change_to_timestamp
from excel_tools.excel_config import TIME_COLUMN, LONG_COLUMN


class ExcelFormat(ExcelUser):
    """
    excel 格式处理类(行高列宽/表黄等处理)
    """

    def __init__(self, excel_name):
        self.excel_name = excel_name
        ExcelUser.__init__(self, excel_name)
        print('[ExcelFormat]path:{}'.format(excel_name))

    def sign(self, sheet_name='Sheet1', height=30, width=20,
             time_column=False, time_range=False):
        """
        :param sheet_name: <path> 表格sheet
        :param time_column:<list> or <str> 表格中时间字段
        :param time_range:<tuple:('2017-03-01', '2019-03-31')> 标定时间范围
                                默认为False，则不需要标黄字段
        :return: a new excel
        """
        columns = []
        if not time_range:
            time_begin = get_first_day_of_month()
            time_end = time.time() * 1000
        if time_range and isinstance(time_range, tuple):
            time_begin = change_to_timestamp(time_range[0])
            time_end = change_to_timestamp(time_range[1])
        data = self.read(sheet_name)
        if time_column and not isinstance(time_column, (str, list)):
            raise (TypeError, 'time_column TypeError')
        if isinstance(time_column, str):
            columns.append(time_column)
        if isinstance(time_column, list):
            columns += time_column
        if data == () or len(data[0]) is None:
            return
        ws = self._wb.get_sheet_by_name(sheet_name)
        ws.row_dimensions[1].value = ''
        index = []
        for column in columns:
            if column not in data[0]:
                continue
            index.append(data[0].index(column))
        font = Font(u'微软雅黑', size=9, color='000000', bold=True)
        font2 = Font(u'微软雅黑', size=9, color='000000')
        bd = Side(style='thin', color="000000")
        fill = PatternFill('solid', fgColor='FFFF00')
        for i in range(1, len(data) + 1):
            ws.row_dimensions[i].height = height
            t = False
            if time_column and i > 1:
                t = ifin_time_range(index, data[i - 1], time_begin, time_end)
            for j in range(1, len(data[0]) + 1):
                ws.cell(row=i, column=j).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                ws.cell(row=i, column=j).value = data[i - 1][j - 1]
                ws.cell(row=i, column=j).font = font2
                ws.cell(row=i, column=j).alignment = Alignment(horizontal='center', vertical='center')
                if i == 1:
                    '首行加粗'
                    ws.cell(row=i, column=j).font = font
                if time_column and t:
                    ws.cell(row=i, column=j).fill = fill
        # excel中字段数量
        column_list = get_cloums_len(len(data[0]))
        for ind, col in enumerate(column_list):
            lens = LONG_COLUMN.get(data[0][ind])
            w = lens if lens else width
            ws.column_dimensions[col].width = w
        name = '{file_name}'.format(file_name=self.excel_name)
        self._wb.save(name)
        print('save successful')
        return True

    def get_all_sheet(self):
        sheets = self._wb.get_sheet_names()
        return sheets


def ifin_time_range(index, data, time_begin, time_end):
    """
    :param index: <list> 对应时间字段 每一行中的下标
    :param data:
    :param time_begin:
    :param time_end:
    :return:
    """
    t = 0
    for d in index:
        if not data[d]:
            continue
        str_time = str(data[d]).strip()
        if not str_time:
            continue
        special_strs = ['None', '0000', '-00-', '-00', '0999-']
        for special_str in special_strs:
            if special_str in str_time:
                return t
        if '.' in str_time:
            str_time = str_time.split('.')[0]
        if ' ' in str_time:
            try:
                stamp = change_to_timestamp(str_time, ft='%Y-%m-%d %H:%M:%S')
            except Exception:
                stamp = change_to_timestamp(str_time, ft='%Y/%m/%d %H:%M:%S')
        if '/' in str_time:
            stamp = change_to_timestamp(str_time, ft='%Y/%m/%d')
        else:
            stamp = change_to_timestamp(str_time.split(' ')[0])
        if stamp >= time_begin and stamp <= (time_end + 68400):
            t = 1
    return t


def get_cloums_len(length):
    """
    :param l: <int> excel 列长度
    :return:<list> excel 相应列名
    """
    colum_length = []
    if length <= 26:
        colum_length = [chr(i).upper() for i in range(97, 97 + length)]
        return colum_length
    else:
        for i in range(97, 97 + 26):
            colum_length.append(chr(i).upper())
        k = 26
        for ii in range(length // 26):
            c_ll = chr(97 + ii).upper()
            for i2 in range(97, 97 + 26):
                k += 1
                if k <= length:
                    c_lll = chr(i2).upper()
                    colum_length.append(c_ll + c_lll)
        return colum_length


def get_all_excel():
    exls = glob.glob('./data/*.xlsx')
    for i in exls:
        yield i


def main():
    time_range = False
    if len(sys.argv) > 2:
        start_time = sys.argv[1:][0]
        end_time = sys.argv[1:][1]
        time_range = (start_time, end_time)
    time_column = TIME_COLUMN
    for excel in get_all_excel():
        print('文件', excel, sep='>')
        exl = ExcelFormat(excel)
        sheets = exl.get_all_sheet()
        for sheet in sheets:
            print('开始', sheet, sep=':')
            # exl.sign(sheet_name=sheet, time_column=time_column, time_range=('2019-10-01', '2019-11-29'))
            exl.sign(sheet_name=sheet, time_column=time_column, time_range=time_range)
    print('格式化表格结束')
    return None


if __name__ == '__main__':
    main()
